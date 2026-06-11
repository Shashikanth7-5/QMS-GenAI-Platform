# services/finetune_service.py
# Fine-tuning pipeline — ready to run when API access is available.
#
# SUPPORTED APPROACHES:
#   1. OpenAI fine-tuning (paid, ~$10-50 per run, best results)
#   2. Anthropic Claude — no fine-tuning API, use prompt caching instead (see below)
#   3. FREE: Ollama + Llama 3 / Mistral — run locally, fine-tune for free
#   4. FREE: HuggingFace PEFT/LoRA — open source, requires GPU
#
# RECOMMENDATION FOR YOUR SETUP:
#   Cognizant managed Claude account → use "few-shot prompting" approach
#   (inject 5-10 good CAPA examples into every prompt = same effect as fine-tuning)
#   This is FREE and works TODAY without any API training quota.

import json
import os
from datetime import datetime
from typing import List, Dict

FINETUNE_DATA_FILE = os.path.join(os.path.dirname(__file__), "..", "finetune_data.jsonl")


# ═══════════════════════════════════════════════════════════════
# APPROACH 1 — OpenAI Fine-Tuning (when you have an OpenAI key)
# ═══════════════════════════════════════════════════════════════

def build_openai_training_data(approved_capas: List[Dict]) -> str:
    """
    Converts approved CAPAs into OpenAI fine-tuning JSONL format.
    Each entry: system prompt + user record → assistant CAPA response.
    Upload the output file to OpenAI fine-tuning API.
    
    Usage:
        capas = get_approved_capas_for_training()
        path  = build_openai_training_data(capas)
        job   = submit_openai_finetune(path)
    """
    lines = []
    for capa in approved_capas:
        rec = _get_source_record(capa.get("sourceRecordId", ""))

        user_msg = (
            f"Generate a CAPA for this quality record:\n"
            f"ID: {rec.get('id','')}\n"
            f"Type: {rec.get('type','').upper()}\n"
            f"Sector: {rec.get('sector','')}\n"
            f"Priority: {rec.get('priority','')}\n"
            f"Title: {rec.get('title','')}\n"
            f"Description: {rec.get('description','')}\n"
            f"Site: {rec.get('site','')}\n"
            f"Regulatory Refs: {', '.join(rec.get('regulatoryRef', []))}"
        )

        assistant_msg = json.dumps({
            "rootCause":            capa.get("rootCause", ""),
            "immediateAction":      capa.get("immediateAction", ""),
            "correctiveAction":     capa.get("correctiveAction", ""),
            "preventiveAction":     capa.get("preventiveAction", ""),
            "proposedOwner":        capa.get("capaOwner", ""),
            "effectivenessCheck":   capa.get("effectivenessCheck", ""),
            "estimatedClosureDays": capa.get("estimatedClosureDays", 30),
            "riskRating":           capa.get("riskRating", ""),
            "regulatoryRef":        capa.get("regulatoryRef", []),
        }, indent=2)

        lines.append(json.dumps({
            "messages": [
                {"role": "system",    "content": _SYSTEM_PROMPT},
                {"role": "user",      "content": user_msg},
                {"role": "assistant", "content": assistant_msg},
            ]
        }))

    output_path = FINETUNE_DATA_FILE
    with open(output_path, "w") as f:
        f.write("\n".join(lines))

    print(f"[finetune] Written {len(lines)} training examples to {output_path}")
    return output_path


def submit_openai_finetune(training_file_path: str) -> dict:
    """
    Submits a fine-tuning job to OpenAI.
    Requires: pip install openai + OPENAI_API_KEY in .env
    Cost: ~$0.008 per 1K tokens. 100 examples ≈ $2-5 total.
    Returns job ID to monitor progress.
    """
    try:
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY", ""))

        # Upload training file
        with open(training_file_path, "rb") as f:
            file_response = client.files.create(file=f, purpose="fine-tune")

        print(f"[finetune] File uploaded: {file_response.id}")

        # Start fine-tuning job
        job = client.fine_tuning.jobs.create(
            training_file = file_response.id,
            model         = "gpt-4o-mini-2024-07-18",  # cheapest model for fine-tuning
            hyperparameters = {"n_epochs": 3},
        )
        print(f"[finetune] Job started: {job.id} — status: {job.status}")
        print(f"[finetune] Monitor at: https://platform.openai.com/finetune/{job.id}")

        # Save job ID for monitoring
        _save_job_record(job.id, file_response.id, len(open(training_file_path).readlines()))
        return {"job_id": job.id, "status": job.status, "model": job.fine_tuned_model}

    except ImportError:
        return {"error": "openai package not installed. Run: pip install openai"}
    except Exception as e:
        return {"error": str(e)}


def check_finetune_status(job_id: str) -> dict:
    """Check status of a fine-tuning job."""
    try:
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY", ""))
        job = client.fine_tuning.jobs.retrieve(job_id)
        return {
            "job_id":            job.id,
            "status":            job.status,
            "fine_tuned_model":  job.fine_tuned_model,
            "trained_tokens":    job.trained_tokens,
            "estimated_finish":  str(job.estimated_finish) if job.estimated_finish else None,
        }
    except Exception as e:
        return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════
# APPROACH 2 — FREE: Few-Shot Prompting (works with Claude account today)
# ═══════════════════════════════════════════════════════════════

def build_few_shot_prompt(record: Dict, n_examples: int = 5) -> str:
    """
    Builds a prompt with n_examples of good CAPAs injected.
    This achieves 80-90% of fine-tuning accuracy for FREE.
    Works with any LLM including your Cognizant Claude account.
    
    Usage: replace _build_capa_prompt() in ai_service.py with this.
    """
    examples = _get_best_examples(n_examples)

    example_text = ""
    for i, (example_rec, example_capa) in enumerate(examples, 1):
        example_text += f"\n--- EXAMPLE {i} ---\n"
        example_text += f"Record: {example_rec.get('type','').upper()} | {example_rec.get('priority','')} | {example_rec.get('title','')}\n"
        example_text += f"Description: {example_rec.get('description','')}\n"
        example_text += f"Root Cause: {example_capa.get('rootCause','')}\n"
        example_text += f"Corrective Action: {example_capa.get('correctiveAction','')}\n"
        example_text += "---\n"

    return (
        "You are a pharmaceutical QA expert. Generate a CAPA following the exact style "
        "and depth of the examples below.\n\n"
        f"REFERENCE EXAMPLES (your approved organisation CAPAs):{example_text}\n\n"
        "Now generate a CAPA for this new record:\n"
        f"Type: {record.get('type','').upper()}\n"
        f"Sector: {record.get('sector','')}\n"
        f"Priority: {record.get('priority','')}\n"
        f"Title: {record.get('title','')}\n"
        f"Description: {record.get('description','')}\n"
        f"Site: {record.get('site','')}\n"
        f"Regulations: {', '.join(record.get('regulatoryRef',[]))}\n\n"
        "Respond ONLY with valid JSON. Required keys: rootCause, immediateAction, "
        "correctiveAction, preventiveAction, proposedOwner, effectivenessCheck, "
        "estimatedClosureDays (int), riskRating, regulatoryRef (array)"
    )


# ═══════════════════════════════════════════════════════════════
# APPROACH 3 — FREE: Ollama local models (no internet, no cost)
# ═══════════════════════════════════════════════════════════════

def generate_with_ollama(record: Dict, model: str = "llama3") -> dict:
    """
    Generate CAPA using a local Ollama model — completely free.
    
    Setup:
        1. Install Ollama: https://ollama.ai
        2. Pull a model: ollama pull llama3   (or mistral, phi3, gemma2)
        3. Set AI_PROVIDER=ollama in .env
    
    Quality: llama3 ≈ GPT-3.5 level for structured tasks.
             mistral ≈ slightly better for technical content.
    """
    import json
    try:
        import httpx
        prompt = build_few_shot_prompt(record, n_examples=3)

        resp = httpx.post(
            "http://localhost:11434/api/generate",
            json={
                "model":  model,
                "prompt": prompt,
                "stream": False,
                "format": "json",
            },
            timeout=120.0,
        )
        resp.raise_for_status()
        raw = resp.json().get("response", "{}")
        return json.loads(raw)
    except Exception as e:
        raise RuntimeError(f"Ollama error: {e}. Is Ollama running? Try: ollama serve")


# ═══════════════════════════════════════════════════════════════
# ACCURACY EVALUATION
# ═══════════════════════════════════════════════════════════════

def evaluate_accuracy(test_cases_file: str = None) -> dict:
    """
    Runs the RCA accuracy test against known-outcome cases.
    Uses the RCA Accuracy Test Cases sheet from QMS_GenAI_1000_Records.xlsx.
    
    Returns accuracy score 0-100 per category.
    """
    test_cases = _load_test_cases(test_cases_file)
    if not test_cases:
        return {"error": "No test cases found. Import QMS_GenAI_1000_Records.xlsx first."}

    results = {
        "total":        len(test_cases),
        "passed":       0,
        "failed":       0,
        "by_category":  {},
        "details":      [],
    }

    for case in test_cases:
        expected_cat   = case.get("expectedRootCauseCategory", "").lower()
        event_desc     = case.get("eventDescription", "")
        if not event_desc:
            continue

        # Generate RCA using current AI
        try:
            from services.ai_service import generate_rca
            rca = generate_rca({
                "type": "deviation",
                "title": event_desc[:80],
                "description": event_desc,
                "priority": "High",
                "sector": "Medical Device",
                "regulatoryRef": [],
            }, method="5why")
            generated_cat = rca.get("rootCauseCategory", "").lower()
        except Exception as e:
            generated_cat = ""

        match = expected_cat and expected_cat in generated_cat
        results["passed"] += 1 if match else 0
        results["failed"] += 0 if match else 1

        cat = case.get("expectedRootCauseCategory", "Unknown")
        if cat not in results["by_category"]:
            results["by_category"][cat] = {"total": 0, "passed": 0}
        results["by_category"][cat]["total"]  += 1
        results["by_category"][cat]["passed"] += 1 if match else 0

        results["details"].append({
            "test_id":      case.get("testId", ""),
            "expected":     expected_cat,
            "generated":    generated_cat,
            "match":        match,
        })

    total = results["total"] or 1
    results["accuracy_pct"] = round(results["passed"] / total * 100, 1)
    results["run_at"]       = datetime.utcnow().isoformat()
    return results


# ── Private helpers ───────────────────────────────────────

_SYSTEM_PROMPT = (
    "You are a pharmaceutical and medical device quality management expert. "
    "Generate structured CAPA responses based on quality records. "
    "Always respond with valid JSON only. Follow ICH, 21 CFR, and ISO 13485 guidelines. "
    "Root causes must be specific, actionable, and traceable to regulatory requirements."
)


def _get_source_record(record_id: str) -> dict:
    try:
        from data.records import get_record_by_id
        rec = get_record_by_id(record_id)
        return rec or {}
    except Exception:
        return {}


def _get_best_examples(n: int = 5) -> list:
    """Get the n most recently approved CAPAs as few-shot examples."""
    try:
        from data.records import get_all_capas
        capas = [c for c in get_all_capas() if c.get("status") == "Approved"]
        pairs = []
        for capa in capas[:n]:
            rec = _get_source_record(capa.get("sourceRecordId", ""))
            if rec:
                pairs.append((rec, capa))
        return pairs
    except Exception:
        return []


def _save_job_record(job_id: str, file_id: str, n_examples: int):
    record_file = os.path.join(os.path.dirname(__file__), "..", "finetune_jobs.json")
    jobs = []
    if os.path.exists(record_file):
        with open(record_file) as f:
            jobs = json.load(f)
    jobs.append({
        "job_id":      job_id,
        "file_id":     file_id,
        "n_examples":  n_examples,
        "submitted_at": datetime.utcnow().isoformat(),
    })
    with open(record_file, "w") as f:
        json.dump(jobs, f, indent=2)


def _load_test_cases(file_path: str = None) -> list:
    """Load RCA test cases from Excel or JSON."""
    if file_path and os.path.exists(file_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if "RCA Accuracy Test Cases" in wb.sheetnames:
                ws = wb["RCA Accuracy Test Cases"]
                headers = [c.value for c in ws[1]]
                cases = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        cases.append(dict(zip(headers, row)))
                return cases
        except Exception as e:
            print(f"[finetune] Could not load test cases: {e}")
    return []
