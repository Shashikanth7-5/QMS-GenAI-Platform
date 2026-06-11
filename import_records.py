# import_records.py
# Run once: python import_records.py
# Imports all 1000 records from QMS_GenAI_1000_Records.xlsx

import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from data.records import add_uploaded_record
from datetime import datetime

EXCEL_PATH = "QMS_GenAI_1000_Records.xlsx"  # place in project root

def run():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: {EXCEL_PATH} not found in project root.")
        print("Place the Excel file in the project root folder.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Quality Records"]

    headers = [c.value for c in ws[1]]
    imported = 0
    skipped  = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        r = dict(zip(headers, row))
        record = {
            "id":            str(r.get("Record ID", "")),
            "type":          str(r.get("Type", "complaint")),
            "sector":        str(r.get("Sector", "Medical Device")),
            "title":         str(r.get("Title", ""))[:90],
            "description":   str(r.get("Description", "")),
            "priority":      str(r.get("Priority", "Medium")),
            "status":        str(r.get("Status", "Draft Generated")),
            "site":          str(r.get("Site", "Unknown")),
            "owner":         str(r.get("Owner", "Unassigned")),
            "detectedDate":  str(r.get("Detected Date", "")),
            "productFamily": str(r.get("Product Family", "")),
            "batchLot":      str(r.get("Batch / Lot", "")),
            "regulatoryRef": str(r.get("Regulatory Refs", "")).split(" | ") if r.get("Regulatory Refs") else [],
            "age":           int(r.get("Age (days)", 0) or 0),
            "createdBy":     str(r.get("Created By", "admin")),
            "createdByName": str(r.get("Created By", "admin")),
            "createdByRole": "admin",
            "_source":       "imported",
        }
        try:
            add_uploaded_record(record)
            imported += 1
            if imported % 100 == 0:
                print(f"  Imported {imported} records...")
        except Exception as e:
            skipped += 1

    print(f"\nDone — {imported} imported, {skipped} skipped")
    print("Restart app.py to see all records in the dashboard.")

if __name__ == "__main__":
    run()